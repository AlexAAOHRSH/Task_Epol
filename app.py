from flask import Flask, jsonify, request, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
from zipfile import ZipFile
from openpyxl.writer.excel import ExcelWriter

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tasks.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Task(db.Model):
	__tablename__ = "Tasks"
	id = db.Column(db.Integer, primary_key=True, autoincrement=True)
	name = db.Column(db.String(100))
	description = db.Column(db.String(500))
	start_date = db.Column(db.DateTime, default=lambda: datetime.today())
	end_date = db.Column(db.DateTime, default=lambda: datetime.today() + timedelta(days=1))
	status = db.Column(db.Integer)
	items = db.relationship("TaskItem", cascade="all,delete")

	def to_dict(self):
		return {
			"id": self.id,
			"name": self.name,
			"description": self.description,
			"start_date": self.start_date.strftime("%Y-%m-%d"),
			"end_date": self.end_date.strftime("%Y-%m-%d"),
			"status": self.status,
			# "items":[
			# 	item.to_dict() for item in db.session.query(TaskItem).filter(TaskItem.task_id == self.id)
			# ]
		}

	@staticmethod
	def validate(body):
		for column in ["name", "description", "start_date", "end_date", "status"]:
			if body.get(column, None) is None:
				raise RuntimeError(column)
		for column in ["start_date", "end_date"]:
			body[column] = datetime.strptime(body.get(column), "%Y-%m-%d")
		body["status"] = int(body["status"])

class TaskItem(db.Model):
	__tablename__ = "TaskItem"

	id = db.Column(db.Integer, primary_key=True)
	name = db.Column(db.String(100))
	description = db.Column(db.String(500))
	value = db.Column(db.String(500))

	task_id = db.Column(db.Integer, db.ForeignKey('Tasks.id'))

	def to_dict(self):
		return {
		"id": self.id,
		"name": self.name,
		"description": self.description,
		"value": self.value,
		"task_id": self.task_id
	}

	@staticmethod
	def validate(body):
		for column in ["name", "description", "value"]:
			if body.get(column, None) is None:
				raise RuntimeError(column)
		body["task_id"] = int(body["task_id"])

db.create_all()

@app.route('/api/tasks', methods=['GET'])
def get_tasks():
	name = request.args.get('name', None)
	status = request.args.get('status', None)
	start_date_from = request.args.get('start_date_from', None)
	end_date_to = request.args.get('end_date_to', None)
	query = db.session.query(Task)

	if name is not None:
		query = query.filter(Task.name.contains(name))
	if status is not None:
		status = int(status)
		query = query.filter(Task.status.contains(status))
	if start_date_from is not None:
		query = query.filter(Task.start_date >= datetime.strptime(start_date_from, "%Y-%m-%d"))
	if end_date_to is not None:
		query = query.filter(Task.end_date <= datetime.strptime(end_date_to, "%Y-%m-%d"))
	return jsonify([task.to_dict() for task in query.all()])
    
@app.route('/api/tasks/items', methods=['GET'])
def get_items():
	value = request.args.get('value', None)
	task_id = request.args.get('task_id', None)
	query = db.session.query(TaskItem)

	if value is not None:
		query = query.filter(TaskItem.value.contains(value))
	if task_id is not None:
		query = query.filter(TaskItem.task_id == task_id)
	return jsonify([task_item.to_dict() for task_item in query.all()])

@app.route('/api/tasks', methods=['POST'])
def  create_task():
	if not request.json:
		return jsonify({"error":"Request is empty"}), 400 
	try:
		Task.validate(request.json)
	except Exception as e:
		return jsonify({"error":f"{e.args[0]} is empty"}), 400

	task = Task(**request.json)
	db.session.add(task)
	db.session.commit()

	return jsonify(task.to_dict()), 201

@app.route('/api/tasks/items', methods=['POST'])
def  create_task_item():
	if not request.json:
		return jsonify({"error":"Request is empty"}), 400 
	try:
		TaskItem.validate(request.json)
	except Exception as e:
		return jsonify({"error":f"{e.args[0]} is empty"}), 400

	task_item = TaskItem(**request.json)
	db.session.add(task_item)
	db.session.commit()

	return jsonify(task_item.to_dict()), 201

@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):

	if not request.json:
		return jsonify({"error":"Request is empty"}), 400 
	try:
		Task.validate(request.json)
	except Exception as e:
		return jsonify({"error":f"{e.args[0]} is empty"}), 400
	db.session.query(Task).filter(Task.id == task_id).update(request.json)
	task = db.session.query(Task).filter(Task.id == task_id).scalar()
	if task is None:
		return jsonify({"error":"Not exists"}), 400
	return jsonify(task.to_dict()), 200

@app.route('/api/tasks/items/<int:task_item_id>', methods=['PUT'])
def update_task_item(task_item_id):

	if not request.json:
		return jsonify({"error":"Request is empty"}), 400 
	try:
		TaskItem.validate(request.json)
	except Exception as e:
		return jsonify({"error":f"{e.args[0]} is empty"}), 400
	db.session.query(TaskItem).filter(TaskItem.id == task_item_id).update(request.json)
	task_item = db.session.query(TaskItem).filter(TaskItem.id == task_item_id).scalar()
	if task_item is None:
		return jsonify({"error":"Not exists"}), 400
	return jsonify(task_item.to_dict()), 200

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
	task = db.session.query(Task).filter(Task.id == task_id).scalar()
	if task is None:
		return jsonify({"error":"Not exists"}), 400
	db.session.delete(task)
	db.session.commit()
	return jsonify({
		**task.to_dict(),
		"message": "This is the last time you see it! :((("
		}), 200

@app.route('/api/tasks/excel', methods=['GET'])
def create_excel():
	wb = Workbook()
	ws = wb.active
	ws.title = "Tasks and items"
	ws[f"{get_column_letter(1)}1"] = "Tasks:"
	ws[f"{get_column_letter(2)}1"] = "id"
	ws[f"{get_column_letter(3)}1"] = "name"
	ws[f"{get_column_letter(4)}1"] = "description"
	ws[f"{get_column_letter(5)}1"] = "start date"
	ws[f"{get_column_letter(6)}1"] = "end date"
	ws[f"{get_column_letter(7)}1"] = "status"
	ws[f"{get_column_letter(8)}1"] = "Items:"
	ws[f"{get_column_letter(9)}1"] = "id"
	ws[f"{get_column_letter(10)}1"] = "name"
	ws[f"{get_column_letter(11)}1"] = "description"
	ws[f"{get_column_letter(12)}1"] = "value"
	ws[f"{get_column_letter(13)}1"] = "task id"

	all_tasks = db.session.query(Task).all()
	rows_offset = 1
	for task in all_tasks:
		ws[f"{get_column_letter(2)}{rows_offset + 1}"] = task.id
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=2,
			end_row=len(task.items)+rows_offset,
			end_column=2
		)
		ws[f"{get_column_letter(3)}{rows_offset + 1}"] = task.name
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=3,
			end_row=len(task.items)+rows_offset,
			end_column=3
		)
		ws[f"{get_column_letter(4)}{rows_offset + 1}"] = task.description
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=4,
			end_row=len(task.items)+rows_offset,
			end_column=4
		)
		ws[f"{get_column_letter(5)}{rows_offset + 1}"] = task.start_date.strftime("%Y-%m-%d")
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=5,
			end_row=len(task.items)+rows_offset,
			end_column=5
		)
		ws[f"{get_column_letter(6)}{rows_offset + 1}"] = task.end_date.strftime("%Y-%m-%d")
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=6,
			end_row=len(task.items)+rows_offset,
			end_column=6
		)
		ws[f"{get_column_letter(7)}{rows_offset + 1}"] = task.status
		ws.merge_cells(
			start_row=rows_offset+1, 
			start_column=7,
			end_row=len(task.items)+rows_offset,
			end_column=7
		)
		all_task_items = task.items
		for i, item in enumerate(all_task_items):
			ws[f"{get_column_letter(9)}{rows_offset + 1 + i}"] = item.id
			ws[f"{get_column_letter(10)}{rows_offset + 1 + i}"] = item.name
			ws[f"{get_column_letter(11)}{rows_offset + 1 + i}"] = item.description
			ws[f"{get_column_letter(12)}{rows_offset + 1 + i}"] = item.value
			ws[f"{get_column_letter(13)}{rows_offset + 1 + i}"] = item.task_id	

		rows_offset += max(1, len(all_task_items))

	content = save_virtual_workbook(wb)
	resp = make_response(content)
	resp.headers["Content-Disposition"] = 'attachment; filename=tasks_info.xlsx'
	resp.headers['Content-Type'] = 'application/x-xlsx'
	return resp

def save_virtual_workbook(workbook,):
    temp_buffer = BytesIO()
    archive = ZipFile(temp_buffer, 'w', ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(workbook, archive)
    try:
        writer.write_data()
    finally:
        archive.close()

    virtual_workbook = temp_buffer.getvalue()
    temp_buffer.close()
    return virtual_workbook	

@app.route('/')
def index():
    return "Hello, World!"

if __name__ == '__main__':
    app.run(debug=True)